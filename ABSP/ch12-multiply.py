# multiplication table

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
for n in range(1,8):
    if n == 1:
        for i in range(2,8):
            sheet.cell(row=n,column=i).value = i-1
        print('row'+str(n))
    else:
        sheet.cell(row=n,column=1).value = n-1
        print('column'+str(n))
        for u in range(2,8):
            sheet.cell(row=n,column=u).value = sheet.cell(row=n,column=1).value*\
            sheet.cell(row=1,column=u).value
wb.save('multiplicationTable.xlsx')
